from flask import Flask, render_template, request
import pandas as pd
import openpyxl
from datetime import datetime

app = Flask(__name__)

GRADE_POINTS = {
    'A+': 10, 'A': 9, 'B': 8, 'C': 7, 'D': 6,
    'E': 5, 'PASS': 0, 'F': 0, 'AB': 0, '': 0, None: 0
}

def calculate_sgpa(subjects):
    total_credits = total_credit_points = 0
    for subj in subjects:
        credits = subj.get("Credits", 0) or 0
        gp = GRADE_POINTS.get(str(subj.get("Performance Grade", "")).strip().upper(), 0)
        subj["Grade Point"] = gp
        subj["Credit Points"] = round(credits * gp, 2)
        total_credits += credits
        total_credit_points += credits * gp
    return round(total_credit_points / total_credits, 2) if total_credits else 0

def extract_semester_data(sheet):
    data, student = [], None
    for row in sheet.iter_rows(min_row=1, values_only=True):
        row = list(row)
        if not any(row): continue
        first_cell = str(row[0]).strip().upper()
        if "NOTE" in first_cell: continue
        if first_cell.startswith("ROLL NO"):
            if student:
                student["SGPA"] = calculate_sgpa(student["Subjects"])
                student["Backlogs"] = sum(1 for s in student["Subjects"] if s["Performance Grade"] == 'F')
                data.append(student)
            student = {"Roll No": str(row[1]).strip() if row[1] else "", "Subjects": []}
        elif first_cell.startswith("NAME") and student:
            student["Name"] = str(row[1]).strip() if row[1] else ""
        elif isinstance(row[0], str) and '-' in row[0] and "Subject Code" not in row[0]:
            try:
                credits = float(row[3]) if row[3] else 0
            except:
                credits = 0
            subject = {
                "Subject Code & Name": row[0],
                "Attendance Grade": row[1],
                "Performance Grade": row[2],
                "Credits": credits
            }
            if student:
                student["Subjects"].append(subject)
    if student:
        student["SGPA"] = calculate_sgpa(student["Subjects"])
        student["Backlogs"] = sum(1 for s in student["Subjects"] if s["Performance Grade"] == 'F')
        data.append(student)
    return data

def extract_all_students(filepath="students_results.xlsx"):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    return {sheet: extract_semester_data(wb[sheet]) for sheet in wb.sheetnames}

def get_student_summary(roll_number, all_data):
    student = {"Roll Number": roll_number, "Name": "", "Semesters": {}, "CGPA": 0, "Backlogs": 0, "Percentage": 0}
    total_credits = total_cp = total_backlogs = 0
    for sem, records in all_data.items():
        for rec in records:
            if rec["Roll No"].strip().upper() == roll_number.strip().upper():
                student["Name"] = rec.get("Name", "Unknown")
                student["Semesters"][sem] = rec
                sem_credits = sum(subj["Credits"] for subj in rec["Subjects"])
                sem_cp = sum(subj["Credit Points"] for subj in rec["Subjects"])
                total_credits += sem_credits
                total_cp += sem_cp
                total_backlogs += rec.get("Backlogs", 0)
    student["CGPA"] = round(total_cp / total_credits, 2) if total_credits else 0
    student["Percentage"] = round(student["CGPA"] * 10 - 5, 2)
    student["Backlogs"] = total_backlogs
    return student if student["Semesters"] else None

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/result', methods=['POST'])
def result():
    roll_no = request.form.get("rollNo", "").strip().upper()
    try:
        all_data = extract_all_students("students_results.xlsx")
        student = get_student_summary(roll_no, all_data)
        if not student:
            return render_template("result.html", error=f"❌ No record found for Roll No: {roll_no}")
        now = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
        return render_template("result.html", student=student, now=now)
    except Exception as e:
        return render_template("result.html", error=f"⚠️ Error: {str(e)}")

if __name__ == "__main__":
    app.run(debug=True)
